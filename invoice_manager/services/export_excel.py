from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from invoice_manager.db import DATA_DIR
from invoice_manager.repositories import (
    add_audit_log,
    list_invoice_allocation_export_rows,
    list_invoices,
    list_work_type_summary,
)
from invoice_manager.utils.date_utils import format_billing_month, validate_billing_month


HEADERS = [
    "請求月",
    "工事コード",
    "工事名",
    "取引先名",
    "担当者名",
    "メールアドレス",
    "電話番号",
    "請求日",
    "請求金額(税込)",
    "添付ファイル数",
    "メモ",
]


def export_monthly_invoice_list(billing_month: str, output_path: Path | None = None) -> Path:
    billing_month = validate_billing_month(billing_month)
    output_path = output_path or _output_path(f"月別請求一覧_{billing_month}.xlsx")
    path = _export_rows("月別請求一覧", list_invoices({"billing_month": billing_month}), output_path)
    add_audit_log("Excel出力", None, None, str(path))
    return path


def export_all_invoice_list(output_path: Path | None = None) -> Path:
    output_path = output_path or _output_path("全件一覧.xlsx")
    path = _export_rows("全件一覧", list_invoices(), output_path)
    add_audit_log("Excel出力", None, None, str(path))
    return path


def export_work_type_summary(output_path: Path | None = None) -> Path:
    output_path = output_path or _output_path("工種コード別集計表.xlsx")
    wb = Workbook()
    for index, (sheet_name, kind) in enumerate(
        [("工事別", "project"), ("取引先別", "vendor"), ("月別", "month")]
    ):
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = sheet_name
        rows = list_work_type_summary(kind)
        _write_table(
            ws,
            ["区分", "工種コード", "工種名", "請求件数", "振分金額"],
            [[row["label"], row["work_type_code"], row["work_type_name"], row["count"], row["total"]] for row in rows],
        )
    path = output_path.resolve()
    wb.save(path)
    add_audit_log("Excel出力", None, None, str(path))
    return path


def export_invoice_allocations(output_path: Path | None = None) -> Path:
    output_path = output_path or _output_path("請求別工種コード振分一覧.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "請求別振分"
    rows = list_invoice_allocation_export_rows()
    _write_table(
        ws,
        [
            "請求ID",
            "請求月",
            "工事コード",
            "工事名",
            "取引先名",
            "請求日",
            "請求金額(税込)",
            "工種コード",
            "工種名",
            "振分金額",
            "振分メモ",
        ],
        [
            [
                row["external_id"],
                format_billing_month(row["billing_month"]),
                row["project_code"],
                row["project_name"],
                row["vendor_name"],
                row["invoice_date"],
                row["total_amount"],
                row["work_type_code"],
                row["work_type_name"],
                row["amount"],
                row["allocation_memo"],
            ]
            for row in rows
        ],
    )
    path = output_path.resolve()
    wb.save(path)
    add_audit_log("Excel出力", None, None, str(path))
    return path


def _output_path(file_name: str) -> Path:
    output_dir = DATA_DIR / "exports"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / file_name


def _export_rows(sheet_name: str, rows: list, output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(
            [
                format_billing_month(row["billing_month"]),
                row["project_code"],
                row["project_name"],
                row["vendor_name"],
                row["contact_name"],
                row["email"],
                row["phone"],
                row["invoice_date"],
                row["total_amount"],
                row["file_count"],
                row["local_memo"],
            ]
        )

    for index, _header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = 18
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["K"].width = 36

    if ws.max_row > 1:
        table_ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
        table = Table(displayName="InvoiceTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
        ws.auto_filter.ref = table_ref

    wb.save(output_path)
    return output_path.resolve()


def _write_table(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(index)].width = max(14, min(36, len(header) + 8))
    if ws.max_row > 1:
        table_ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        table = Table(displayName=f"Table{abs(hash(ws.title))}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
        ws.auto_filter.ref = table_ref
