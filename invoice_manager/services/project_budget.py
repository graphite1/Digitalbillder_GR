"""Project budget storage, source review, and provisional cost forecasts.

Source documents are deliberately kept separate from saved budget rows.  A PDF or
workbook preview can suggest cell values, but callers must explicitly confirm the
reviewed/edited rows before this module writes a budget.
"""
from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
from dataclasses import asdict, dataclass, replace
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Mapping

from invoice_manager import db
from invoice_manager.services.work_type_resolution import (
    CanonicalWorkType,
    WorkTypeResolutionError,
    load_work_type_choices,
    resolve_from_catalog,
)


SUPPORTED_SOURCE_SUFFIXES = frozenset({".pdf", ".xlsx", ".xlsm"})
MAX_SOURCE_BYTES = 50 * 1024 * 1024
_MONEY_RE = re.compile(r"^[+-]?(?:\d{1,3}(?:,\d{3})*|\d+)$")
_HEADER_WORDS = {
    "コード", "工種コード", "科目", "実行予算", "予定金額", "備考",
    "code", "item", "budget", "scheduled", "remarks",
}


@dataclass(frozen=True, slots=True)
class BudgetRowInput:
    work_type_code: str
    work_type_name: str
    budget_net: int
    remaining_net: int | None = None
    scheduled_net: int | None = None
    include_in_total: bool = False
    actual_work_type_code: str | None = None
    source_candidate: "ExtractedBudgetCandidate | None" = None
    row_id: int | None = None


@dataclass(frozen=True, slots=True)
class BudgetRow:
    id: int
    work_type_code: str
    work_type_name: str
    budget_net: int
    remaining_net: int | None
    scheduled_net: int | None
    include_in_total: bool
    actual_work_type_code: str | None
    source_candidate_json: str | None
    source_proposal_id: int | None
    edit_version: int
    edited_at: str
    sort_order: int

    @property
    def source_candidate(self) -> dict[str, object] | None:
        return json.loads(self.source_candidate_json) if self.source_candidate_json else None


@dataclass(frozen=True, slots=True)
class ProjectBudget:
    id: int
    project_id: int
    source_original_name: str | None
    source_stored_path: str | None
    source_sha256: str | None
    source_type: str | None
    note: str
    source_proposal_id: int | None
    updated_at: str
    rows: tuple[BudgetRow, ...]

    @property
    def total_budget_net(self) -> int:
        return sum(row.budget_net for row in self.rows if row.include_in_total)


@dataclass(frozen=True, slots=True)
class ExtractedBudgetCandidate:
    """Untrusted source value which can only be used after UI review."""

    page_number: int
    work_type_code: str
    work_type_name: str
    budget_net: int | None
    scheduled_net: int | None
    source_location: str
    aggregation_hint: str = "要確認"
    requires_review: bool = True


@dataclass(frozen=True, slots=True)
class SourcePreview:
    path: Path
    source_type: str
    page_count: int | None
    preview_text: str
    candidates: tuple[ExtractedBudgetCandidate, ...]
    warnings: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class ForecastRow:
    work_type_code: str
    work_type_name: str
    budget_net: int
    actual_net: int | None
    remaining_net: int | None
    projected_final_net: int | None
    variance_net: int | None
    include_in_total: bool
    actual_work_type_code: str | None
    is_unmapped_actual: bool


_SCHEMA = (
    """
    CREATE TABLE IF NOT EXISTS project_budgets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        project_id INTEGER NOT NULL UNIQUE,
        source_original_name TEXT,
        source_stored_path TEXT,
        source_sha256 TEXT,
        source_type TEXT,
        note TEXT NOT NULL DEFAULT '',
        source_proposal_id INTEGER,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY(project_id) REFERENCES projects(id)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS project_budget_rows (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        budget_id INTEGER NOT NULL,
        work_type_code TEXT NOT NULL,
        work_type_name TEXT NOT NULL,
        budget_net INTEGER NOT NULL CHECK(budget_net >= 0),
        remaining_net INTEGER CHECK(remaining_net IS NULL OR remaining_net >= 0),
        scheduled_net INTEGER CHECK(scheduled_net IS NULL OR scheduled_net >= 0),
        include_in_total INTEGER NOT NULL DEFAULT 0 CHECK(include_in_total IN (0, 1)),
        actual_work_type_code TEXT,
        source_candidate_json TEXT,
        source_proposal_id INTEGER,
        edit_version INTEGER NOT NULL DEFAULT 1,
        edited_at TEXT NOT NULL,
        sort_order INTEGER NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY(budget_id) REFERENCES project_budgets(id) ON DELETE CASCADE,
        FOREIGN KEY(source_proposal_id) REFERENCES project_budget_source_proposals(id),
        UNIQUE(budget_id, work_type_code)
    )
    """,
    "CREATE INDEX IF NOT EXISTS idx_project_budget_rows_code ON project_budget_rows(work_type_code)",
    """
    CREATE TABLE IF NOT EXISTS project_budget_source_proposals (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        project_id INTEGER NOT NULL,
        source_sha256 TEXT NOT NULL,
        source_original_name TEXT NOT NULL,
        source_type TEXT NOT NULL,
        page_number INTEGER,
        proposal_json TEXT NOT NULL,
        created_at TEXT NOT NULL,
        UNIQUE(project_id, source_sha256, page_number),
        FOREIGN KEY(project_id) REFERENCES projects(id)
    )
    """,
)


def _ensure_schema(connection) -> None:
    for statement in _SCHEMA:
        connection.execute(statement)
    # These tables are owned by this module.  Keep early local builds readable.
    budget_columns = {row["name"] for row in connection.execute("PRAGMA table_info(project_budgets)")}
    if "source_proposal_id" not in budget_columns:
        connection.execute("ALTER TABLE project_budgets ADD COLUMN source_proposal_id INTEGER")
    row_columns = {row["name"] for row in connection.execute("PRAGMA table_info(project_budget_rows)")}
    additions = {
        "scheduled_net": "ALTER TABLE project_budget_rows ADD COLUMN scheduled_net INTEGER",
        "include_in_total": "ALTER TABLE project_budget_rows ADD COLUMN include_in_total INTEGER NOT NULL DEFAULT 0",
        "actual_work_type_code": "ALTER TABLE project_budget_rows ADD COLUMN actual_work_type_code TEXT",
        "source_candidate_json": "ALTER TABLE project_budget_rows ADD COLUMN source_candidate_json TEXT",
        "source_proposal_id": "ALTER TABLE project_budget_rows ADD COLUMN source_proposal_id INTEGER",
        "edit_version": "ALTER TABLE project_budget_rows ADD COLUMN edit_version INTEGER NOT NULL DEFAULT 1",
        "edited_at": "ALTER TABLE project_budget_rows ADD COLUMN edited_at TEXT NOT NULL DEFAULT ''",
    }
    for name, statement in additions.items():
        if name not in row_columns:
            connection.execute(statement)


def initialize_project_budgets() -> None:
    with db.get_connection() as connection:
        _ensure_schema(connection)


def _now_text() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _required_text(value: object, label: str) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        raise ValueError(f"{label}は必須です。")
    return text


def _nonnegative_integer(value: object, label: str, *, optional: bool = False) -> int | None:
    if value is None and optional:
        return None
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        raise ValueError(f"{label}は0以上の整数で指定してください。")
    return value


def _normalize_row(row: BudgetRowInput | Mapping[str, object]) -> BudgetRowInput:
    if isinstance(row, Mapping):
        row = BudgetRowInput(
            work_type_code=str(row.get("work_type_code", "")),
            work_type_name=str(row.get("work_type_name", "")),
            budget_net=row.get("budget_net"),  # type: ignore[arg-type]
            remaining_net=row.get("remaining_net"),  # type: ignore[arg-type]
            scheduled_net=row.get("scheduled_net"),  # type: ignore[arg-type]
            include_in_total=row.get("include_in_total", False) is True,
            actual_work_type_code=(None if row.get("actual_work_type_code") is None else str(row.get("actual_work_type_code"))),
            source_candidate=row.get("source_candidate"),  # type: ignore[arg-type]
            row_id=row.get("row_id"),  # type: ignore[arg-type]
        )
    if not isinstance(row, BudgetRowInput):
        raise TypeError("rowsにはBudgetRowInputまたは同じキーを持つ辞書を指定してください。")
    if row.source_candidate is not None and not isinstance(row.source_candidate, ExtractedBudgetCandidate):
        raise TypeError("source_candidateには抽出候補を指定してください。")
    row_id = row.row_id
    if row_id is not None and (isinstance(row_id, bool) or not isinstance(row_id, int) or row_id <= 0):
        raise ValueError("予算行IDが不正です。")
    return BudgetRowInput(
        _required_text(row.work_type_code, "工種コード"),
        _required_text(row.work_type_name, "工種名"),
        _nonnegative_integer(row.budget_net, "実行予算(税抜)"),  # type: ignore[arg-type]
        _nonnegative_integer(row.remaining_net, "残工事見込(税抜)", optional=True),
        _nonnegative_integer(row.scheduled_net, "予定金額(税抜)", optional=True),
        bool(row.include_in_total),
        None if row.actual_work_type_code is None else (str(row.actual_work_type_code).strip() or None),
        row.source_candidate,
        row_id,
    )


def prepare_budget_rows_from_candidates(
    candidates: Iterable[ExtractedBudgetCandidate],
    *,
    existing_codes: Iterable[str] = (),
    fallback_names: Mapping[str, str] | None = None,
    project_id: int | None = None,
    existing_actual_codes: Iterable[str] = (),
) -> tuple[tuple[BudgetRowInput, ...], tuple[str, ...]]:
    """Validate a complete preview and prepare review-only budget editor rows.

    Existing editor rows are protected by skipping matching codes.  Validation is
    deliberately completed before any result is returned, so a malformed or
    duplicate source proposal cannot lead to a partial bulk addition.
    """

    source_candidates = tuple(candidates)
    protected_codes = {
        _required_text(code, "既存工種コード") for code in existing_codes
    }
    name_fallbacks = fallback_names or {}
    catalog = load_work_type_choices(project_id) if project_id is not None else ()
    prepared: list[BudgetRowInput] = []
    seen_locations: dict[str, str] = {}

    for index, candidate in enumerate(source_candidates, start=1):
        if not isinstance(candidate, ExtractedBudgetCandidate):
            raise ValueError(f"抽出候補{index}の形式が不正です。")
        location = candidate.source_location.strip() or f"候補{index}"
        try:
            code = _required_text(candidate.work_type_code, "工種コード")
        except ValueError as exc:
            raise ValueError(f"抽出候補 {location}: {exc}") from exc
        if code in seen_locations:
            raise ValueError(
                f"抽出候補の工種コードが重複しています: {code} "
                f"（{seen_locations[code]} / {location}）"
            )
        seen_locations[code] = location

        candidate_name = str(candidate.work_type_name or "").strip()
        fallback_name = str(name_fallbacks.get(code, "") or "").strip()
        if not candidate_name and not fallback_name and catalog:
            try:
                fallback_name = resolve_from_catalog(code, catalog).name.strip()
            except WorkTypeResolutionError:
                pass
        name = candidate_name or fallback_name
        if not name:
            raise ValueError(f"抽出候補 {location}（{code}）: 工種名は必須です。")

        row = BudgetRowInput(
            work_type_code=code,
            work_type_name=name,
            budget_net=candidate.budget_net,  # type: ignore[arg-type]
            remaining_net=None,
            scheduled_net=candidate.scheduled_net,
            include_in_total=False,
            actual_work_type_code=None,
            source_candidate=candidate,
        )
        try:
            prepared.append(_normalize_row(row))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"抽出候補 {location}（{code}）: {exc}") from exc

    rows = tuple(row for row in prepared if row.work_type_code not in protected_codes)
    skipped = tuple(row.work_type_code for row in prepared if row.work_type_code in protected_codes)
    if project_id is not None:
        rows, _ = _suggest_budget_mappings(rows, catalog, existing_actual_codes)
    return rows, skipped


def _suggest_budget_mappings(
    rows: tuple[BudgetRowInput, ...],
    catalog: tuple[CanonicalWorkType, ...],
    reserved_actual_codes: Iterable[str] = (),
) -> tuple[tuple[BudgetRowInput, ...], tuple[str, ...]]:
    """Offer unique mappings without changing printed codes or explicit mappings."""
    reserved = set(reserved_actual_codes) | {
        row.actual_work_type_code for row in rows if row.actual_work_type_code
    }
    proposals: dict[int, str] = {}
    issues: list[str] = []
    for index, row in enumerate(rows):
        if row.actual_work_type_code:
            continue
        try:
            proposals[index] = resolve_from_catalog(row.work_type_code, catalog).code
        except WorkTypeResolutionError as exc:
            issues.append(f"{row.work_type_code}: {exc}")
    counts: dict[str, int] = {}
    for code in proposals.values():
        counts[code] = counts.get(code, 0) + 1
    result = list(rows)
    for index, code in proposals.items():
        if code in reserved or counts[code] > 1:
            issues.append(f"{rows[index].work_type_code}: Web工種 {code} は他の予算行と重複するため未対応です。")
        else:
            result[index] = replace(rows[index], actual_work_type_code=code)
    return tuple(result), tuple(issues)


def suggest_budget_work_type_mappings(
    project_id: int,
    rows: Iterable[BudgetRowInput | Mapping[str, object]],
) -> tuple[tuple[BudgetRowInput, ...], tuple[str, ...]]:
    """Prepare unsaved mappings for review; never replace explicit saved choices."""
    normalized = tuple(_normalize_row(row) for row in rows)
    return _suggest_budget_mappings(normalized, load_work_type_choices(project_id))


def _normalize_rows(rows: Iterable[BudgetRowInput | Mapping[str, object]]) -> tuple[BudgetRowInput, ...]:
    normalized = tuple(_normalize_row(row) for row in rows)
    if not normalized:
        raise ValueError("予算行を1件以上入力してください。")
    seen: set[str] = set()
    duplicates: list[str] = []
    for row in normalized:
        if row.work_type_code in seen:
            duplicates.append(row.work_type_code)
        seen.add(row.work_type_code)
    if duplicates:
        raise ValueError(f"工種コードが重複しています: {', '.join(dict.fromkeys(duplicates))}")
    mappings = [row.actual_work_type_code for row in normalized if row.actual_work_type_code]
    duplicate_mappings = [code for code in mappings if mappings.count(code) > 1]
    if duplicate_mappings:
        raise ValueError(
            "同じWeb工種コードを複数行へ対応付けできません: "
            + ", ".join(dict.fromkeys(duplicate_mappings))
        )
    row_ids = [row.row_id for row in normalized if row.row_id is not None]
    duplicate_row_ids = [row_id for row_id in row_ids if row_ids.count(row_id) > 1]
    if duplicate_row_ids:
        raise ValueError(
            "同じ予算行IDが重複しています: "
            + ", ".join(str(value) for value in dict.fromkeys(duplicate_row_ids))
        )
    if not any(row.include_in_total for row in normalized):
        raise ValueError("集計対象の予算行を1件以上選択してください。")
    return normalized


def _validate_source(path: str | Path) -> Path:
    source = Path(path).expanduser().resolve(strict=True)
    if not source.is_file():
        raise ValueError("予算原本にはファイルを指定してください。")
    if source.suffix.lower() not in SUPPORTED_SOURCE_SUFFIXES:
        raise ValueError("予算原本はPDF、.xlsx、.xlsmに対応しています。")
    if source.stat().st_size > MAX_SOURCE_BYTES:
        raise ValueError("予算原本は50MB以下にしてください。")
    return source


def _is_under(path: Path, parent: Path) -> bool:
    try:
        return os.path.commonpath((str(path), str(parent))) == str(parent)
    except ValueError:
        return False


def resolve_budget_source(stored_path: str) -> Path:
    """Resolve a stored relative path while rejecting traversal and absolute paths."""
    relative = Path(stored_path)
    if relative.is_absolute():
        raise ValueError("予算原本の保存先が不正です。")
    data_root = db.DATA_DIR.resolve()
    resolved = (data_root / relative).resolve()
    if not _is_under(resolved, data_root):
        raise ValueError("予算原本の保存先がデータ領域外です。")
    return resolved


def _copy_source(source_path: str | Path, project_id: int) -> tuple[str, str, str, str]:
    source = _validate_source(source_path)
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    safe_name = re.sub(r"[^\w.()-]+", "_", source.name, flags=re.UNICODE).strip("._")
    safe_name = safe_name[-100:] or f"source{source.suffix.lower()}"
    relative = Path("budgets") / f"project-{int(project_id)}" / f"{digest[:16]}_{safe_name}"
    destination = resolve_budget_source(relative.as_posix())
    destination.parent.mkdir(parents=True, exist_ok=True)
    if not destination.exists():
        shutil.copy2(source, destination)
    return source.name, relative.as_posix(), digest, source.suffix.lower().lstrip(".")


def save_project_budget(
    project_id: int,
    rows: Iterable[BudgetRowInput | Mapping[str, object]],
    *,
    source_path: str | Path | None = None,
    source_preview: SourcePreview | None = None,
    confirmed: bool = False,
    note: str = "",
) -> ProjectBudget:
    """Replace one project's reviewed budget rows atomically.

    ``confirmed=True`` is intentionally mandatory.  Preview candidates are never
    passed here automatically; a caller must first let the user select or edit rows.
    """
    if not confirmed:
        raise ValueError("原本の候補を確認し、登録する予算行を選択・編集してください。")
    if isinstance(project_id, bool) or not isinstance(project_id, int) or project_id <= 0:
        raise ValueError("工事IDが不正です。")
    normalized = _normalize_rows(rows)
    with db.get_connection() as connection:
        project_exists = connection.execute("SELECT 1 FROM projects WHERE id = ?", (project_id,)).fetchone()
    if project_exists is None:
        raise ValueError("工事が見つかりません。")
    if source_preview is not None and source_path is None:
        source_path = source_preview.path
    if source_preview is not None and Path(source_path).resolve() != source_preview.path.resolve():
        raise ValueError("確認した原本と保存する原本が一致しません。")
    source_values = _copy_source(source_path, project_id) if source_path is not None else None
    timestamp = _now_text()

    with db.atomic_transaction():
        with db.get_connection() as connection:
            _ensure_schema(connection)
            project = connection.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone()
            if project is None:
                raise ValueError("工事が見つかりません。")
            proposal_id: int | None = None
            if source_preview is not None and source_values is not None:
                proposal_json = json.dumps(
                    [asdict(candidate) for candidate in source_preview.candidates],
                    ensure_ascii=False, sort_keys=True,
                )
                page_number = source_preview.candidates[0].page_number if source_preview.candidates else None
                connection.execute(
                    """
                    INSERT OR IGNORE INTO project_budget_source_proposals (
                        project_id, source_sha256, source_original_name, source_type,
                        page_number, proposal_json, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (project_id, source_values[2], source_values[0], source_values[3],
                     page_number, proposal_json, timestamp),
                )
                proposal = connection.execute(
                    """
                    SELECT id FROM project_budget_source_proposals
                    WHERE project_id = ? AND source_sha256 = ? AND page_number IS ?
                    """,
                    (project_id, source_values[2], page_number),
                ).fetchone()
                proposal_id = int(proposal["id"])
            existing = connection.execute(
                "SELECT * FROM project_budgets WHERE project_id = ?", (project_id,)
            ).fetchone()
            if existing is None:
                source_values = source_values or (None, None, None, None)
                cursor = connection.execute(
                    """
                    INSERT INTO project_budgets (
                        project_id, source_original_name, source_stored_path,
                        source_sha256, source_type, note, source_proposal_id, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (project_id, *source_values, note.strip(), proposal_id, timestamp, timestamp),
                )
                budget_id = int(cursor.lastrowid)
            else:
                budget_id = int(existing["id"])
                if source_values is None:
                    source_values = (
                        existing["source_original_name"], existing["source_stored_path"],
                        existing["source_sha256"], existing["source_type"],
                    )
                connection.execute(
                    """
                    UPDATE project_budgets
                    SET source_original_name = ?, source_stored_path = ?, source_sha256 = ?,
                        source_type = ?, note = ?, source_proposal_id = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (*source_values, note.strip(),
                     proposal_id if proposal_id is not None else existing["source_proposal_id"],
                     timestamp, budget_id),
                )
            existing_rows_by_code = {
                str(item["work_type_code"]): item
                for item in connection.execute(
                    "SELECT * FROM project_budget_rows WHERE budget_id = ?", (budget_id,)
                ).fetchall()
            }
            existing_rows_by_id = {
                int(item["id"]): item for item in existing_rows_by_code.values()
            }
            # Release the per-budget code uniqueness before applying row-ID based
            # edits.  This permits code swaps while the surrounding transaction
            # keeps the temporary identifiers invisible to other readers.
            for existing_id in existing_rows_by_id:
                connection.execute(
                    "UPDATE project_budget_rows SET work_type_code = ? WHERE id = ?",
                    (f"\x1fediting:{budget_id}:{existing_id}:{timestamp}", existing_id),
                )
            kept_ids: set[int] = set()
            claimed_existing_ids: set[int] = set()
            for index, row in enumerate(normalized, start=1):
                if row.row_id is not None:
                    previous = existing_rows_by_id.get(row.row_id)
                    if previous is None:
                        raise ValueError("指定された予算行IDはこの工事に存在しません。再読み込みしてください。")
                else:
                    previous = existing_rows_by_code.get(row.work_type_code)
                    if previous is not None and int(previous["id"]) in claimed_existing_ids:
                        previous = None
                if previous is not None:
                    claimed_existing_ids.add(int(previous["id"]))
                source_json = (
                    json.dumps(asdict(row.source_candidate), ensure_ascii=False, sort_keys=True)
                    if row.source_candidate is not None
                    else (previous["source_candidate_json"] if previous is not None else None)
                )
                row_proposal_id = (
                    proposal_id if row.source_candidate is not None and proposal_id is not None
                    else (previous["source_proposal_id"] if previous is not None else None)
                )
                if previous is None:
                    connection.execute(
                        """
                        INSERT INTO project_budget_rows (
                            budget_id, work_type_code, work_type_name, budget_net,
                            remaining_net, scheduled_net, include_in_total,
                            actual_work_type_code, source_candidate_json, edit_version,
                            source_proposal_id, edited_at, sort_order, created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?)
                        """,
                        (budget_id, row.work_type_code, row.work_type_name, row.budget_net,
                         row.remaining_net, row.scheduled_net, int(row.include_in_total),
                         row.actual_work_type_code, source_json, row_proposal_id,
                         timestamp, index, timestamp, timestamp),
                    )
                else:
                    kept_ids.add(int(previous["id"]))
                    connection.execute(
                        """
                        UPDATE project_budget_rows
                        SET work_type_code = ?, work_type_name = ?, budget_net = ?, remaining_net = ?, scheduled_net = ?,
                            include_in_total = ?, actual_work_type_code = ?, source_candidate_json = ?,
                            source_proposal_id = ?, edit_version = edit_version + 1,
                            edited_at = ?, sort_order = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (row.work_type_code, row.work_type_name, row.budget_net, row.remaining_net, row.scheduled_net,
                         int(row.include_in_total), row.actual_work_type_code, source_json,
                         row_proposal_id, timestamp, index, timestamp, int(previous["id"])),
                    )
            for existing_id in existing_rows_by_id:
                if existing_id not in kept_ids:
                    connection.execute("DELETE FROM project_budget_rows WHERE id = ?", (existing_id,))
    result = get_project_budget(project_id)
    if result is None:  # pragma: no cover - defensive guard after a committed insert
        raise RuntimeError("保存した予算を読み込めませんでした。")
    return result


def get_project_budget(project_id: int) -> ProjectBudget | None:
    with db.get_connection() as connection:
        _ensure_schema(connection)
        budget = connection.execute(
            "SELECT * FROM project_budgets WHERE project_id = ?", (int(project_id),)
        ).fetchone()
        if budget is None:
            return None
        rows = connection.execute(
            "SELECT * FROM project_budget_rows WHERE budget_id = ? ORDER BY sort_order, id",
            (int(budget["id"]),),
        ).fetchall()
    return ProjectBudget(
        id=int(budget["id"]), project_id=int(budget["project_id"]),
        source_original_name=budget["source_original_name"],
        source_stored_path=budget["source_stored_path"],
        source_sha256=budget["source_sha256"], source_type=budget["source_type"],
        note=str(budget["note"] or ""),
        source_proposal_id=(None if budget["source_proposal_id"] is None else int(budget["source_proposal_id"])),
        updated_at=str(budget["updated_at"]),
        rows=tuple(
            BudgetRow(int(row["id"]), str(row["work_type_code"]), str(row["work_type_name"]),
                      int(row["budget_net"]),
                      None if row["remaining_net"] is None else int(row["remaining_net"]),
                      None if row["scheduled_net"] is None else int(row["scheduled_net"]),
                      bool(row["include_in_total"]), row["actual_work_type_code"],
                      row["source_candidate_json"],
                      None if row["source_proposal_id"] is None else int(row["source_proposal_id"]),
                      int(row["edit_version"]),
                      str(row["edited_at"]), int(row["sort_order"]))
            for row in rows
        ),
    )


def list_source_proposals(project_id: int) -> tuple[dict[str, object], ...]:
    """Return immutable extraction proposals, newest first, for audit/comparison."""
    with db.get_connection() as connection:
        _ensure_schema(connection)
        rows = connection.execute(
            """
            SELECT id, source_sha256, source_original_name, source_type,
                   page_number, proposal_json, created_at
            FROM project_budget_source_proposals
            WHERE project_id = ? ORDER BY id DESC
            """,
            (int(project_id),),
        ).fetchall()
    return tuple(
        {
            "id": int(row["id"]), "source_sha256": str(row["source_sha256"]),
            "source_original_name": str(row["source_original_name"]),
            "source_type": str(row["source_type"]), "page_number": row["page_number"],
            "candidates": json.loads(str(row["proposal_json"])), "created_at": str(row["created_at"]),
        }
        for row in rows
    )


def build_project_forecast(project_id: int) -> tuple[ForecastRow, ...]:
    budget = get_project_budget(project_id)
    if budget is None:
        return ()
    from invoice_manager.services.historical_costs import get_historical_sync_status, list_actual_costs

    actual_by_code: dict[str, int] = {}
    actual_names: dict[str, str] = {}
    actuals_are_known = get_historical_sync_status().last_successful_refresh is not None
    if actuals_are_known:
        for item in list_actual_costs(project_id):
            code = str(item.work_type_code)
            actual_by_code[code] = actual_by_code.get(code, 0) + int(item.net_amount)
            actual_names.setdefault(code, str(item.work_type_name))

    budget_by_code = {row.work_type_code: row for row in budget.rows}
    mapped_actual_codes = {
        row.actual_work_type_code for row in budget.rows if row.actual_work_type_code is not None
    }
    ordered_codes = [row.work_type_code for row in budget.rows]
    unmatched_actual_codes = sorted(set(actual_by_code) - mapped_actual_codes)
    result = []
    for code in ordered_codes:
        row = budget_by_code.get(code)
        budget_net = row.budget_net if row else 0
        remaining = row.remaining_net if row else None
        actual_code = row.actual_work_type_code if row else code
        actual = actual_by_code.get(actual_code, 0) if actuals_are_known and actual_code else None
        projected = None if remaining is None or actual is None else actual + remaining
        result.append(ForecastRow(
            code, row.work_type_name if row else actual_names.get(code, code), budget_net,
            actual, remaining, projected,
            None if projected is None else budget_net - projected,
            row.include_in_total if row else False,
            actual_code,
            False,
        ))
    for code in unmatched_actual_codes:
        result.append(ForecastRow(
            code, f"{actual_names.get(code, code)}（予算行との対応未設定）", 0,
            actual_by_code[code], None, None, None, False, code, True,
        ))
    return tuple(result)


def preview_source_document(path: str | Path, *, page_number: int = 2) -> SourcePreview:
    source = _validate_source(path)
    if source.suffix.lower() == ".pdf":
        return _preview_pdf(source, page_number=page_number)
    return _preview_workbook(source)


def _parse_money(value: object) -> int | None:
    text = "" if value is None else str(value).strip().replace("¥", "").replace("￥", "")
    if not text or not _MONEY_RE.fullmatch(text):
        return None
    return int(text.replace(",", ""))


def _looks_like_code(value: object) -> bool:
    text = "" if value is None else str(value).strip()
    if not text or len(text) > 32 or any(character.isspace() for character in text):
        return False
    if "�" in text or text.casefold() in _HEADER_WORDS:
        return False
    return not text.endswith("%")


def _budget_column_groups(rows: list[list[object]]) -> tuple[list[tuple[int, int, int, int]], bool]:
    """Infer code/name/budget/scheduled columns from adjacent numeric cell pairs."""
    for row in rows:
        labels = [re.sub(r"\s+", "", "" if value is None else str(value)) for value in row]
        budget_columns = [index for index, label in enumerate(labels) if label == "実行予算"]
        groups = []
        for budget_column in budget_columns:
            nearby_scheduled = next(
                (index for index in range(budget_column + 1, min(len(labels), budget_column + 4))
                 if labels[index] == "予定金額"),
                None,
            )
            if nearby_scheduled is None:
                continue
            code_column = next(
                (index for index in range(budget_column - 1, -1, -1)
                 if labels[index] in {"コード", "工種コード"}),
                budget_column - 2,
            )
            name_column = next(
                (index for index in range(code_column + 1, budget_column)
                 if labels[index] == "科目"),
                budget_column - 1,
            )
            groups.append((code_column, name_column, budget_column, nearby_scheduled))
        if groups:
            return groups, False
    max_columns = max((len(row) for row in rows), default=0)
    groups = []
    for budget_column in range(2, max_columns - 1):
        score = 0
        for row in rows:
            if len(row) <= budget_column + 1:
                continue
            if _parse_money(row[budget_column]) is not None and _parse_money(row[budget_column + 1]) is not None:
                score += 1
        if score >= 2:
            candidate = (budget_column - 2, budget_column - 1, budget_column, budget_column + 1)
            if not groups or candidate[0] > groups[-1][3]:
                groups.append(candidate)
    return groups, True


def _candidates_from_table(
    rows: list[list[object]], page_number: int, table_number: int
) -> tuple[list[ExtractedBudgetCandidate], bool]:
    result = []
    groups, structurally_inferred = _budget_column_groups(rows)
    for code_column, name_column, budget_column, scheduled_column in groups:
        for row_number, row in enumerate(rows, start=1):
            padded = list(row) + [None] * max(0, scheduled_column + 1 - len(row))
            code = "" if padded[code_column] is None else str(padded[code_column]).strip()
            if not _looks_like_code(code):
                continue
            budget_value = _parse_money(padded[budget_column])
            scheduled_value = _parse_money(padded[scheduled_column])
            # A blank-amount row is still useful for a printed code such as 513.
            # Restrict that case to compact identifier characters so adjacent free
            # text (material/work descriptions) cannot become a code candidate.
            if budget_value is None and scheduled_value is None and not re.fullmatch(r"[A-Za-z0-9_.:/-]+", code):
                continue
            name = "" if padded[name_column] is None else str(padded[name_column]).strip()
            if "�" in name:
                name = ""
            result.append(ExtractedBudgetCandidate(
                page_number, code, name, budget_value,
                scheduled_value, f"表{table_number} 行{row_number}",
            ))
    return result, structurally_inferred


def _preview_pdf(source: Path, *, page_number: int) -> SourcePreview:
    if page_number <= 0:
        raise ValueError("PDFのページ番号は1以上で指定してください。")
    try:
        import pymupdf
    except ImportError:  # pragma: no cover - compatibility with older PyMuPDF
        import fitz as pymupdf  # type: ignore[no-redef]

    warnings: list[str] = []
    candidates: list[ExtractedBudgetCandidate] = []
    used_structural_inference = False
    with pymupdf.open(source) as document:
        page_count = len(document)
        if page_number > page_count:
            raise ValueError(f"PDFは{page_count}ページです。{page_number}ページ目はありません。")
        page = document[page_number - 1]
        raw_text = page.get_text("text", sort=True).strip()
        try:
            tables = page.find_tables().tables
            for table_number, table in enumerate(tables, start=1):
                table_candidates, inferred = _candidates_from_table(
                    table.extract(), page_number, table_number
                )
                candidates.extend(table_candidates)
                used_structural_inference = used_structural_inference or inferred
        except Exception as exc:
            warnings.append(f"罫線表を読み取れませんでした: {exc}")

    duplicate_codes: list[str] = []
    seen_codes: set[str] = set()
    for candidate in candidates:
        if candidate.work_type_code in seen_codes:
            duplicate_codes.append(candidate.work_type_code)
        seen_codes.add(candidate.work_type_code)
    if duplicate_codes:
        warnings.append(
            "同じコードが複数箇所にあります。候補は統合していません。登録する行を選んでください: "
            + ", ".join(dict.fromkeys(duplicate_codes))
        )
    if used_structural_inference:
        warnings.append(
            "見出し名を特定できない表は数値列の並びから構造を推定しました。実行予算と予定金額の列対応を確認してください。"
        )
    if raw_text.count("�") > max(3, len(raw_text) // 20):
        warnings.append("PDFの日本語文字層を正しく復元できません。名称は工種マスタまたは原本画像で確認してください。")
    if not candidates:
        warnings.append("予算候補を抽出できません。画像PDFや未知の表形式は手入力してください。")
    warnings.append("抽出値は確認用候補です。選択・修正した行だけが登録対象になります。")
    preview_lines = [
        f"{item.work_type_code}\t{item.work_type_name or '（名称要確認）'}\t"
        f"{'' if item.budget_net is None else format(item.budget_net, ',')}\t"
        f"{'' if item.scheduled_net is None else format(item.scheduled_net, ',')}"
        for item in candidates
    ]
    preview_text = "コード\t科目\t実行予算\t予定金額\n" + "\n".join(preview_lines)
    return SourcePreview(source, "pdf", page_count, preview_text, tuple(candidates), tuple(warnings))


def _preview_workbook(source: Path) -> SourcePreview:
    from openpyxl import load_workbook

    workbook = load_workbook(source, read_only=True, data_only=True)
    lines: list[str] = []
    try:
        for sheet in workbook.worksheets[:3]:
            lines.append(f"[{sheet.title}]")
            for row in sheet.iter_rows(min_row=1, max_row=80, max_col=20, values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(values):
                    lines.append("\t".join(values).rstrip())
    finally:
        workbook.close()
    return SourcePreview(
        source, source.suffix.lower().lstrip("."), None, "\n".join(lines)[:20000], (),
        ("Excelは内容プレビューのみです。原本を確認し、予算行を手入力してください。",),
    )
